import datetime
import os

import openpyxl
from aiogram import Router, F
from aiogram.fsm.context import FSMContext
from aiogram.types import CallbackQuery, FSInputFile

import FSM
import utils
from polling_manager import add_bot, stop_bot
from dispatchers import dp_new_bot, polling_manager
from db_ import db
import markup.admin_markup as admin_markup


callbacks_router = Router()


@callbacks_router.callback_query(F.data.startswith("answer_feedback_"))
async def answer_feedback_(callback: CallbackQuery, state: FSMContext):
    message_id = callback.data.split("_")[-1]
    if not db.check_answer_message(message_id):
        await callback.message.answer("Введите текст для ответа пользователю", reply_markup=admin_markup.create_cancel_markup())
        await state.set_state(FSM.FSMAdmin.get_text_answer_user)
        await state.set_data({"message_id": message_id})
    else:
        await callback.message.answer("На данный вопрос уже был дан ответ")


@callbacks_router.callback_query(F.data.startswith("ban_"))
async def ban_user(callback: CallbackQuery, state: FSMContext):
    bot_id = callback.data.split("_")[-1]
    user_id = callback.data.split("_")[-1]
    db.ban_user(user_id, bot_id)
    await callback.message.answer("Пользователь был успешно заблокирован!")


@callbacks_router.callback_query(F.data.startswith("edit_marks"))
async def edit_marks(callback: CallbackQuery):
    bot_id = callback.data.split("_")[-1]
    marks = db.get_marks(bot_id)
    await callback.message.edit_text("Настройка системы UTM меток\n\nВыберите UTM метку ниже или создайте новую",
                                  reply_markup=admin_markup.create_markup_choice_marks(bot_id, marks))


@callbacks_router.callback_query(F.data.startswith("edit_start_message_"))
async def edit_start_message_(callback: CallbackQuery, state: FSMContext):
    bot_id = callback.data.split("_")[-1]
    await callback.message.answer("Теперь пришлите мне текст стартового сообщения для персонажа", reply_markup=admin_markup.create_cancel_markup())
    await state.set_state(FSM.FSMAdmin.get_new_start_text)
    await state.set_data({"bot_id": bot_id})


@callbacks_router.callback_query(F.data.startswith("mark"))
async def show_mark(callback: CallbackQuery):
    mark_id = callback.data.split("_")[-2]
    bot_id = callback.data.split("_")[-1]
    mark = db.get_mark_by_id(mark_id)
    if len(mark) > 0:
        mark_name = mark[0][2]
        mark_link = mark[0][3]
        count_conversion = db.get_conversion_by_mark(mark_id)
        await callback.message.edit_text(f"<b>UTM метка</b>\n\nНазвание: {mark_name}\nСсылка: {mark_link}\nПереходов: {count_conversion}",
                                      parse_mode="HTML", reply_markup=admin_markup.create_markup_mark(mark_id, bot_id))
    else:
        await callback.message.answer("Данная метка была удалена!")


@callbacks_router.callback_query(F.data.startswith("table_users_mark_"))
async def table_users_mark_(callback: CallbackQuery, state: FSMContext):
    mark_id = callback.data.split("_")[-1]
    titles = ["User ID", f"Username", "Полное имя", "Дата регистрации", "Сумма покупок"]
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    for i, value in enumerate(titles, 1):
        ws.cell(row=1, column=i).value = value

    users = {}
    users_with_payments, all_users = db.get_users_by_mark_id(mark_id)
    for user in all_users:
        if not users.get(user[0], False):
            users[user[0]] = [user[1], user[2], user[3], 0]
    for user in users_with_payments:
        if not users.get(user[0], False):
            users[user[0]] = [user[1], user[2], user[3], user[4]]
        users[user[0]][3] += user[4]

    for index, user in enumerate(users, 2):
        row = [user] + users[user]
        for i, value in enumerate(row, 1):
            ws.cell(row=index, column=i).value = value

    path = f"Таблица пользователей.xlsx"
    workbook.save(path)
    await callback.message.answer_document(document=FSInputFile(path))
    os.remove(path)


@callbacks_router.callback_query(F.data.startswith("table_marks_"))
async def table_marks_(callback: CallbackQuery, state: FSMContext):
    bot_id = callback.data.split("_")[-1]
    marks = db.get_marks(bot_id)
    titles = ["Название", f"Идентификатор", "Кол-во переходов", "Сумма покупок"]
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    for i, value in enumerate(titles, 1):
        ws.cell(row=1, column=i).value = value

    for index, mark in enumerate(marks, 2):
        count_conversion = db.get_conversion_by_mark(mark[0])
        count_buy = db.get_count_buy_by_mark(mark[0])
        if count_buy is None:
            count_buy = 0
        row = [mark[2], mark[3].split("=")[-1], count_conversion, count_buy]
        for i, value in enumerate(row, 1):
            ws.cell(row=index, column=i).value = value

    path = f"Таблица меток.xlsx"
    workbook.save(path)
    await callback.message.answer_document(document=FSInputFile(path))
    os.remove(path)


@callbacks_router.callback_query(F.data.startswith("create_mark_"))
async def create_mark_(callback: CallbackQuery, state: FSMContext):
    bot_id = callback.data.split("_")[-1]
    await state.set_state(FSM.FSMAdmin.get_mark_name)
    await state.set_data({"bot_id": bot_id})
    await callback.message.answer("Пришлите мне наименование метки")


@callbacks_router.callback_query(F.data.startswith("back_to_settings_"))
async def back_to_settings_(callback: CallbackQuery):
    bot_id = callback.data.split("_")[-1]

    bots = db.get_bots()
    msg = ""
    status = ""
    for bot in bots:
        if str(bot[0]) == bot_id:
            status = bot[3]
            msg = f"[{bot[3]}] - @{bot[2]}\n"
            break
    await callback.message.edit_text(msg, reply_markup=admin_markup.create_markup_start_stop_bot(status, bot_id))


@callbacks_router.callback_query(F.data.startswith("delete_mark_"))
async def delete_mark_(callback: CallbackQuery):
    bot_id = callback.data.split("_")[-1]
    mark_id = callback.data.split("_")[-2]
    db.delete_mark(mark_id)
    marks = db.get_marks(bot_id)
    await callback.message.edit_text("Успешно удалил метку!\n\nНастройка системы UTM меток\n\nВыберите UTM метку ниже или создайте новую",
                                  reply_markup=admin_markup.create_markup_choice_marks(bot_id, marks))


@callbacks_router.callback_query(F.data.startswith("manual_pay_"))
async def manual_pay_(callback: CallbackQuery):
    bot_id = callback.data.split("_")[-1]
    pay_types = db.get_pay_types(bot_id)
    await callback.message.edit_text("Настройка способов оплаты", reply_markup=admin_markup.create_markup_manual_pay(bot_id, pay_types))


@callbacks_router.callback_query(F.data.startswith("statistics_manual_type_"))
async def statistics_manual_type_(callback: CallbackQuery):
    bot_id = callback.data.split("_")[-1]
    pay_types_in_db = db.get_pay_types(bot_id)
    pay_types = {"yoomoney": [0, 0]}
    for pay_type in pay_types_in_db:
        pay_types[pay_type[2]] = [0, 0]

    payments = db.get_payments_for_statistics(bot_id)
    for payment in payments:
        if not pay_types.get(payment[1], False):
            pay_types[payment[1]] = [0, 0]
        pay_types[payment[1]][0] += 1
        pay_types[payment[1]][1] += payment[0]
    titles = ["Название", f"Кол-во оплат", "Сумма покупок"]
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    for i, value in enumerate(titles, 1):
        ws.cell(row=1, column=i).value = value

    for index, pay_type in enumerate(pay_types, 2):
        row = [pay_type, pay_types[pay_type][0], pay_types[pay_type][1]]
        for i, value in enumerate(row, 1):
            ws.cell(row=index, column=i).value = value

    path = f"Статистика оплат.xlsx"
    workbook.save(path)
    await callback.message.answer_document(document=FSInputFile(path))
    os.remove(path)


@callbacks_router.callback_query(F.data.startswith("add_manual_type_"))
async def add_manual_type_(callback: CallbackQuery, state: FSMContext):
    bot_id = callback.data.split("_")[-1]
    await state.set_state(FSM.FSMAdmin.get_title_pay_type)
    await state.set_data({"bot_id": bot_id})
    await callback.message.answer("Пришлите мне наименование способа оплаты", reply_markup=admin_markup.create_cancel_markup())


@callbacks_router.callback_query(F.data.startswith("manual_type_"))
async def manual_type_(callback: CallbackQuery, state: FSMContext):
    pay_type_id = callback.data.split("_")[-1]
    pay_type = db.get_pay_type_by_id(pay_type_id)
    await callback.message.edit_text(f"<b>Название</b>: {pay_type[2]}\n\n<b>Описание</b>: {pay_type[3]}",
                                     reply_markup=admin_markup.create_markup_edit_manual_pay_type(pay_type_id, pay_type[1], pay_type[4]))


@callbacks_router.callback_query(F.data.startswith("back_to_manual_pay_"))
async def back_to_manual_pay_(callback: CallbackQuery):
    bot_id = callback.data.split("_")[-1]
    pay_types = db.get_pay_types(bot_id)
    await callback.message.edit_text("Настройка способов оплаты", reply_markup=admin_markup.create_markup_manual_pay(bot_id, pay_types))


@callbacks_router.callback_query(F.data.startswith("change_name_pay_type_"))
async def change_name_pay_type_(callback: CallbackQuery, state: FSMContext):
    pay_type_id = callback.data.split("_")[-1]
    await state.set_state(FSM.FSMAdmin.get_new_title_pay_type)
    await state.set_data({"pay_id": pay_type_id})
    await callback.message.answer("Пришлите мне новое название способа оплаты", reply_markup=admin_markup.create_cancel_markup())


@callbacks_router.callback_query(F.data.startswith("change_description_pay_type_"))
async def change_description_pay_type_(callback: CallbackQuery, state: FSMContext):
    pay_type_id = callback.data.split("_")[-1]
    await state.set_state(FSM.FSMAdmin.get_new_description_pay_type)
    await state.set_data({"pay_id": pay_type_id})
    await callback.message.answer("Пришлите мне новое описание способа оплаты",
                                  reply_markup=admin_markup.create_cancel_markup())


@callbacks_router.callback_query(F.data.startswith("disable_pay_type_"))
async def disable_pay_type_(callback: CallbackQuery):
    pay_type_id = callback.data.split("_")[-1]
    db.update_status_pay_type(pay_type_id)
    pay_type = db.get_pay_type_by_id(pay_type_id)
    await callback.message.edit_text(f"<b>Название</b>: {pay_type[2]}\n\n<b>Описание</b>: {pay_type[3]}",
                                     reply_markup=admin_markup.create_markup_edit_manual_pay_type(pay_type_id, pay_type[1], pay_type[4]))


@callbacks_router.callback_query(F.data.startswith("delete_pay"))
async def delete_pay_type_(callback: CallbackQuery):
    pay_type_id = callback.data.split("_")[-1]
    db.delete_pay_type(pay_type_id)
    bot_id = callback.data.split("_")[-2]
    pay_types = db.get_pay_types(bot_id)
    await callback.message.edit_text("<b>Успешно удалил способ оплаты</b>\n\nНастройка способов оплаты",
                                     reply_markup=admin_markup.create_markup_manual_pay(bot_id, pay_types))


@callbacks_router.callback_query(F.data.startswith("back_to_mark_settings_"))
async def back_to_mark_settings_(callback: CallbackQuery):
    bot_id = callback.data.split("_")[-1]
    marks = db.get_marks(bot_id)
    await callback.message.edit_text("Настройка системы UTM меток\n\nВыберите UTM метку ниже или создайте новую",
                                  reply_markup=admin_markup.create_markup_choice_marks(bot_id, marks))


@callbacks_router.callback_query(F.data == "add_photo")
async def add_photo(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Пришлите мне картинку (не забудь поставить галочку на пункте «Сжать изображение»)")
    await state.set_state(FSM.FSMAdmin.get_start_photo)


@callbacks_router.callback_query(F.data == "skip_add_photo")
async def skip_add_photo(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    bot_id = data["token"].split(":")[0]
    await state.set_state(FSM.FSMAdmin.get_rates)
    await callback.message.answer("Теперь давайте настроим тарифы", reply_markup=admin_markup.create_markup_rates(bot_id))


@callbacks_router.callback_query(F.data == "edit_photo")
async def edit_photo(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Пришлите мне картинку (не забудь поставить галочку на пункте «Сжать изображение»)")
    await state.set_state(FSM.FSMAdmin.get_new_start_photo)


@callbacks_router.callback_query(F.data == "skip_edit_photo")
async def skip_edit_photo(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    db.update_start_photo(data["bot_id"], None)
    await callback.message.answer("Успешно обновил стартовое сообщение", reply_markup=admin_markup.create_start_markup())
    await state.clear()


@callbacks_router.callback_query(F.data == "next_settings_rate")
async def next_settings_rate(callback: CallbackQuery, state: FSMContext):

    data = await state.get_data()
    rates = db.get_all_rates_by_bot_id(data["token"].split(":")[0])
    if len(rates) > 0:
        await state.clear()
        await callback.message.answer("Главное меню", reply_markup=admin_markup.create_start_markup())
    else:

        await callback.message.answer("Добавьте хотя бы один тариф, чтобы закончить настройку")


@callbacks_router.callback_query(F.data.startswith("edit_main_rates"))
async def edit_rates(callback: CallbackQuery, state: FSMContext):
    await state.set_state(FSM.FSMAdmin.get_rates)
    bot_id = callback.data.split("_")[-1]
    token = db.get_bot_token(bot_id)
    await state.set_data({"token": token})
    rates = db.get_all_rates_by_bot_id(bot_id)
    current_rates = ""
    for rate in rates:
        current_rates += f"{rate[2]} минут - {rate[3]}₽ / {rate[4]}$\n"
    await callback.message.answer(f"Текущие тарифы:\n{current_rates}",
                         reply_markup=admin_markup.create_markup_rates(bot_id))


@callbacks_router.callback_query(F.data.startswith("edit_rates"))
async def edit_rates(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    bot_id = data["token"].split(":")[0]
    rates = db.get_all_rates_by_bot_id(bot_id)
    await callback.message.edit_text("Выберите тариф для изменения", reply_markup=admin_markup.create_markup_edit_rate(rates))


@callbacks_router.callback_query(F.data.startswith("edit_rate"))
async def edit_rate(callback: CallbackQuery, state: FSMContext):
    rate_id = callback.data.split("_")[-1]
    await callback.message.delete()
    data = await state.get_data()
    msg = await callback.message.answer("Введите новое время в минутах тарифа (целое число)")
    data["edit_rate_id"] = rate_id
    data["msg_id"] = msg.message_id

    await state.set_data(data)
    await state.set_state(FSM.FSMAdmin.get_new_count_minutes)


@callbacks_router.callback_query(F.data.startswith("add_rate"))
async def add_rate(callback: CallbackQuery, state: FSMContext):
    await state.set_state(FSM.FSMAdmin.get_count_minutes)
    await callback.message.delete()
    msg = await callback.message.answer("Введите время в минутах тарифа (целое число)", reply_markup=admin_markup.create_markup_rate_back())
    data = await state.get_data()
    data["msg_id"] = msg.message_id
    await state.set_data(data)


@callbacks_router.callback_query(F.data.startswith("delete_rates"))
async def delete_rates(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    bot_id = data["token"].split(":")[0]
    rates = db.get_all_rates_by_bot_id(bot_id)
    await callback.message.edit_text("Выберите тариф для удаления", reply_markup=admin_markup.create_markup_delete_rate(rates))


@callbacks_router.callback_query(F.data == "rate_back")
async def rate_back(callback: CallbackQuery, state: FSMContext):

    data = await state.get_data()
    bot_id = data["token"].split(":")[0]
    rates = db.get_all_rates_by_bot_id(bot_id)
    current_rates = ""
    for rate in rates:
        current_rates += f"{rate[2]} минут - {rate[3]}₽ / {rate[4]}$\n"
    await callback.message.delete()
    await callback.message.answer(f"Текущие тарифы:\n{current_rates}",
                         reply_markup=admin_markup.create_markup_rates(bot_id))


@callbacks_router.callback_query(F.data.startswith("delete_rate_"))
async def delete_rate(callback: CallbackQuery, state: FSMContext):
    rate_id = callback.data.split("_")[-1]
    db.delete_rate(rate_id)
    data = await state.get_data()
    bot_id = data["token"].split(":")[0]
    rates = db.get_all_rates_by_bot_id(bot_id)
    current_rates = ""
    for rate in rates:
        current_rates += f"{rate[2]} минут - {rate[3]}₽ / {rate[4]}$\n"
    await callback.message.edit_text(f"Успешно удалил тариф!\n\nТекущие тарифы:\n{current_rates}",
                         reply_markup=admin_markup.create_markup_rates(bot_id))


@callbacks_router.callback_query(F.data.startswith("start_bot"))
async def start_bot_callback(callback: CallbackQuery):
    bot_id = callback.data.split("_")[-1]
    bot_token = db.get_bot_token(bot_id)
    _, _, bot_id = await add_bot(bot_token, dp_new_bot, polling_manager)
    if bot_id:
        db.update_status_bot(bot_token, "запущен")
        # await callback.message.answer("Персонаж успешно запущен!", reply_markup=admin_markup.create_start_markup())
        await callback.message.edit_text(text=f'<i>Персонаж успешно запущен!</i>\n\n{callback.message.text[callback.message.text.index("["):].replace("остановлен", "запущен").replace("не запущен", "запущен")}', reply_markup=admin_markup.create_markup_start_stop_bot('запущен', bot_id), parse_mode='HTML')
    else:
        await callback.message.answer("Не удалось запустить персонажа", reply_markup=admin_markup.create_start_markup())


@callbacks_router.callback_query(F.data.startswith("stop_bot"))
async def stop_bot_callback(callback: CallbackQuery):
    bot_id = int(callback.data.split("_")[-1])
    bot_token = db.get_bot_token(bot_id)
    if await stop_bot(bot_id, polling_manager):
        db.update_status_bot(bot_token, "остановлен")
        # await callback.message.answer("Персонаж остановлен", reply_markup=admin_markup.create_start_markup())
        await callback.message.edit_text(text=f'<i>Персонаж остановлен</i>\n\n{callback.message.text[callback.message.text.index("["):].replace("запущен", "остановлен")}', reply_markup=admin_markup.create_markup_start_stop_bot('остановлен', bot_id), parse_mode='HTML')
    else:
        await callback.message.answer("Не удалось остановить персонажа", reply_markup=admin_markup.create_start_markup())


@callbacks_router.callback_query(F.data.startswith("promocodes"))
async def promocodes(callback: CallbackQuery):
    bot_id = callback.data.split("_")[-1]
    promocodes = db.get_promocodes(bot_id)
    await callback.message.edit_text("<b>Настройки промокодов</b>\n\nВыберите нужный раздел промокодов или создайте новый", reply_markup=admin_markup.create_markup_promocodes(promocodes, bot_id))


@callbacks_router.callback_query(F.data.startswith("add_promocode"))
async def add_promocode(callback: CallbackQuery, state: FSMContext):
    bot_id = callback.data.split("_")[-1]
    await state.set_state(FSM.FSMAdmin.get_title_promo)
    await state.set_data({"bot_id": bot_id})
    await callback.message.answer("Пришлите мне наименование промокода", reply_markup=admin_markup.create_cancel_markup())


@callbacks_router.callback_query(F.data.startswith("back_to_promocodes_"))
async def back_to_promocodes_(callback: CallbackQuery):
    bot_id = callback.data.split("_")[-1]
    promocodes = db.get_promocodes(bot_id)
    await callback.message.edit_text("<b>Настройки промокодов</b>\n\nВыберите нужный раздел промокодов или создайте новый", reply_markup=admin_markup.create_markup_promocodes(promocodes, bot_id))


@callbacks_router.callback_query(F.data.startswith("delete_promo_"))
async def delete_promo_(callback: CallbackQuery):
    promo_id = callback.data.split("_")[-2]
    bot_id = callback.data.split("_")[-1]
    db.delete_promo(promo_id)
    promocodes = db.get_promocodes(bot_id)
    await callback.message.edit_text("<b>Успешно удалил промокод!\n\nНастройки промокодов</b>\n\nВыберите нужный раздел промокодов или создайте новый", reply_markup=admin_markup.create_markup_promocodes(promocodes, bot_id))


@callbacks_router.callback_query(F.data.startswith("edit_promo"))
async def edit_count_activation_total_(callback: CallbackQuery, state: FSMContext):
    bot_id = callback.data.split("_")[-1]
    promo_id = callback.data.split("_")[-2]
    type_action = callback.data.split("_")[-3]
    if type_action == "total":
        await state.set_state(FSM.FSMAdmin.get_new_count_activation_total)
        await callback.message.answer("Пришлите мне новое количество активаций промокода", reply_markup=admin_markup.create_cancel_markup())
    elif type_action == "person":
        await state.set_state(FSM.FSMAdmin.get_new_count_activation_by_person)
        await callback.message.answer("Пришлите мне новое количество активаций промокода на человека", reply_markup=admin_markup.create_cancel_markup())
    elif type_action == "date":
        await state.set_state(FSM.FSMAdmin.get_new_date_end)
        await callback.message.answer("Пришлите мне новую дату окончания действия промокода в формате ГГГГ-ММ-ДД ЧЧ:ММ", reply_markup=admin_markup.create_cancel_markup())

    await state.set_data({"promo_id": promo_id, "bot_id": bot_id})


@callbacks_router.callback_query(F.data.startswith("promocode"))
async def show_promo(callback: CallbackQuery):
    bot_id = callback.data.split("_")[-2]
    promo_id = callback.data.split("_")[-1]
    msg = utils.create_message_promo(promo_id)
    await callback.message.edit_text(msg, reply_markup=admin_markup.create_markup_promocode(bot_id, promo_id))


@callbacks_router.callback_query(F.data.startswith("promo_statistics_"))
async def promo_statistics_(callback: CallbackQuery):
    promo_id = callback.data.split("_")[-1]

    titles = ["User ID", f"Username", "Полное имя", "Дата регистрации", "UTM метка", "UTM название", "Сумма покупок", "Кол-во активаций"]
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    for i, value in enumerate(titles, 1):
        ws.cell(row=1, column=i).value = value

    users = db.get_users_by_promo_id(promo_id)
    bot_id = db.get_promo(promo_id)[0][1]
    marks = db.get_marks(bot_id)
    users_result = {}
    for user in users:
        if not users_result.get(user[0], False):
            utm_mark = ["", ""]
            for mark in marks:
                if mark[0] == user[5]:
                    utm_mark = [mark[3].split("=")[-1], mark[2]]
                    break
            users_result[user[0]] = [user[1], user[2], user[3], utm_mark[0], utm_mark[1], user[4], 1]
        else:
            users_result[user[0]][5] += user[4]
            users_result[user[0]][6] += 1

    for index, user in enumerate(users_result, 2):
        row = [user] + users_result[user]
        for i, value in enumerate(row, 1):
            ws.cell(row=index, column=i).value = value

    path = f"Список активаций.xlsx"
    workbook.save(path)
    await callback.message.answer_document(document=FSInputFile(path))
    os.remove(path)


@callbacks_router.callback_query(F.data.startswith("mailing"))
async def mailing(callback: CallbackQuery, state: FSMContext):
    bot_id = callback.data.split("_")[-1]
    await state.set_state(FSM.FSMAdmin.get_text)
    await state.set_data({"bot_id": bot_id})
    await callback.message.answer("Пришлите мне текст для рассылки")


@callbacks_router.callback_query(F.data.startswith("edit_prompt_"))
async def edit_prompt_(callback: CallbackQuery, state: FSMContext):
    bot_id = int(callback.data.split("_")[-1])
    await callback.message.delete()
    await callback.message.answer("Пришлите мне новый промпт", reply_markup=admin_markup.create_cancel_markup())
    await state.set_state(FSM.FSMAdmin.get_new_prompt)
    await state.set_data({"bot_id": bot_id})


@callbacks_router.callback_query(F.data.startswith("edit_yoomoney_token_"))
async def edit_yoomoney_token(callback: CallbackQuery, state: FSMContext):
    bot_id = int(callback.data.split("_")[-1])
    await callback.message.delete()
    await callback.message.answer("Пришлите мне новый токен yoomoney", reply_markup=admin_markup.create_cancel_markup())
    await state.set_state(FSM.FSMAdmin.get_new_yoomoney_token)
    await state.set_data({"bot_id": bot_id})


@callbacks_router.callback_query(F.data.startswith("edit_voice_"))
async def edit_voice_(callback: CallbackQuery, state: FSMContext):
    bot_id = int(callback.data.split("_")[-1])
    await callback.message.delete()
    await callback.message.answer("Пришлите мне новый голос персонажа", reply_markup=admin_markup.create_cancel_markup())
    await state.set_state(FSM.FSMAdmin.get_new_voice)
    await state.set_data({"bot_id": bot_id})



@callbacks_router.callback_query(F.data.startswith("edit_price_per_minute_"))
async def edit_price_per_minute_(callback: CallbackQuery, state: FSMContext):
    bot_id = int(callback.data.split("_")[-1])
    await callback.message.delete()
    await callback.message.answer("Пришлите мне новую цену за минуту персонажа", reply_markup=admin_markup.create_cancel_markup())
    await state.set_state(FSM.FSMAdmin.get_new_price_per_minute)
    await state.set_data({"bot_id": bot_id})



@callbacks_router.callback_query(F.data.startswith("statistics"))
async def statistics(callback: CallbackQuery, state: FSMContext):
    bot_id = callback.data.split("_")[-1]
    payments = db.get_payments(bot_id)
    count_users = db.get_count_users(bot_id)
    count_users_last_30_days = db.get_count_users_last_30_days(bot_id)

    total_subscribes = 0
    total_subscribe_current_month = 0
    total_subscribe_last_30_days = 0
    total_subscribe_last_24_hours = 0
    total_subscribe_today = 0

    total_profit_last_24_hour = 0
    total_profit_today = 0
    total_profit_yesterday = 0
    total_profit_current_month = 0
    total_profit_last_30_days = 0
    total_profit = 0

    for payment in payments:
        payment_datetime = datetime.datetime.strptime(payment[0], "%Y-%m-%d %H:%M")
        current_date = datetime.datetime.now()
        total_subscribes += 1
        total_profit += payment[1]

        if payment_datetime.month == current_date.month:
            total_subscribe_current_month += 1
            total_profit_current_month += payment[1]
        if payment_datetime + datetime.timedelta(days=30) > current_date:
            total_subscribe_last_30_days += 1
            total_profit_last_30_days += payment[1]
        if payment_datetime + datetime.timedelta(hours=24) > current_date:
            total_subscribe_last_24_hours += 1
            total_profit_last_24_hour += payment[1]
        if payment_datetime.strftime("%Y-%m-%d") == current_date.strftime("%Y-%m-%d"):
            total_subscribe_today += 1
            total_profit_today += payment[1]
        if payment_datetime.strftime("%Y-%m-%d") == (current_date - datetime.timedelta(days=1)).strftime("%Y-%m-%d"):
            total_profit_yesterday += payment[1]

    average_price = 0
    if total_subscribes > 0:
        average_price = round(total_profit / total_subscribes)

    await callback.message.answer(f"""
📊<b>Статистика вашего бота</b>

Всего переходов в бота: {count_users} чел.
Переходов в бота за последние 30 дней: {count_users_last_30_days} чел.

👥<b>Количество подписчиков:</b>
🔸Купили подписку за всё время: {total_subscribes} раз.
🔸Купили подписку за текущий месяц: {total_subscribe_current_month} раз.
🔸Купили подписку за 30 дней: {total_subscribe_last_30_days} раз.
🔸Купили подписку за 24 часа: {total_subscribe_last_24_hours} раз.
🔸Купили подписку за сегодня: {total_subscribe_today} раз.

💰<b>Доход проекта за:</b>
🔷24 часа - {total_profit_last_24_hour}₽.
🔷сегодня - {total_profit_today}₽.
🔷вчера - {total_profit_yesterday}₽.
🔷текущий месяц - {total_profit_current_month}₽.
🔷за последние 30 дней - {total_profit_last_30_days}₽.
🔷за всё время - {total_profit}₽.
🔷средний чек - {average_price}₽.
""", parse_mode="HTML")
